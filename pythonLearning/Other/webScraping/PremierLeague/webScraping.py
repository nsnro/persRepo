import os
import sys
import time
from datetime import date

import openpyxl
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

## link = https://www.premierleague.com/stats/top/clubs/wins

links = ["https://www.premierleague.com/stats/top/clubs/wins?se=1",
               "https://www.premierleague.com/stats/top/clubs/wins?se=2",
               "https://www.premierleague.com/stats/top/clubs/wins?se=3",
               "https://www.premierleague.com/stats/top/clubs/wins?se=4",
               "https://www.premierleague.com/stats/top/clubs/wins?se=5",
               "https://www.premierleague.com/stats/top/clubs/wins?se=6",
               "https://www.premierleague.com/stats/top/clubs/wins?se=7",
               "https://www.premierleague.com/stats/top/clubs/wins?se=8",
               "https://www.premierleague.com/stats/top/clubs/wins?se=9",
               "https://www.premierleague.com/stats/top/clubs/wins?se=10",
               "https://www.premierleague.com/stats/top/clubs/wins?se=11",
               "https://www.premierleague.com/stats/top/clubs/wins?se=12",
               "https://www.premierleague.com/stats/top/clubs/wins?se=13",
               "https://www.premierleague.com/stats/top/clubs/wins?se=14",
               "https://www.premierleague.com/stats/top/clubs/wins?se=15",
               "https://www.premierleague.com/stats/top/clubs/wins?se=16",
               "https://www.premierleague.com/stats/top/clubs/wins?se=17",
               "https://www.premierleague.com/stats/top/clubs/wins?se=18",
               "https://www.premierleague.com/stats/top/clubs/wins?se=19",
               "https://www.premierleague.com/stats/top/clubs/wins?se=20",
               "https://www.premierleague.com/stats/top/clubs/wins?se=21",
               "https://www.premierleague.com/stats/top/clubs/wins?se=22",
               "https://www.premierleague.com/stats/top/clubs/wins?se=27",
               "https://www.premierleague.com/stats/top/clubs/wins?se=42",
               "https://www.premierleague.com/stats/top/clubs/wins?se=54",
               "https://www.premierleague.com/stats/top/clubs/wins?se=79",
               "https://www.premierleague.com/stats/top/clubs/wins?se=210",
               "https://www.premierleague.com/stats/top/clubs/wins?se=274",
               "https://www.premierleague.com/stats/top/clubs/wins?se=363",
               "https://www.premierleague.com/stats/top/clubs/wins?se=418",
               "https://www.premierleague.com/stats/top/clubs/wins?se=489",
               "https://www.premierleague.com/stats/top/clubs/wins?se=578"]


def getData(browser, path):
    finalTable = []
    finalWins = []
    seasonYear = browser.find_element(By.XPATH, '/html/body/main/div[2]/div[1]/section/div[2]/div[2]')
    print(f"Season year: {seasonYear.get_attribute("innerText")}")

    gatheringDone = False

    while not gatheringDone:
        # teamNames = []
        # teamWins = []
        teamNames = browser.find_elements(By.CLASS_NAME, 'stats-table__cell-icon-align')
        for names in teamNames:
            finalTable.append(names.text)

        teamWins = browser.find_elements(By.CLASS_NAME, 'stats-table__main-stat')
        for wins in teamWins:
            finalWins.append(wins.text)

        try:
            browser.find_element(By.CSS_SELECTOR,
                                 "#mainContent > div.wrapper.col-12 > div:nth-child(1) > div.paginationContainer > div.paginationBtn.paginationNextContainer.inactive")
        except NoSuchElementException:
            button = browser.find_element(By.CSS_SELECTOR,
                                          "#mainContent > div.wrapper.col-12 > div:nth-child(1) > div.paginationContainer > div.paginationBtn.paginationNextContainer")
            button.click()
            time.sleep(5)
        else:
            gatheringDone = True
            continue

    fillData(seasonYear, finalTable, finalWins, path)


def fillData(year, teamName, teamWins, path):
    wb = openpyxl.load_workbook(path)
    yearSheet = year.get_attribute("innerText")
    yearFinal = yearSheet.replace("/", "-")
    sheet = wb.create_sheet(title=str(yearFinal))

    for i in range(len(teamName)):
        data = (str(teamName[i]), int(teamWins[i]))
        sheet.append(data)

    wb.save(path)
    print("Data inserted")


def precondition():
    if sys.argv[1] is None:
        print("Please enter a path")
        return
    else:
        if not os.path.isdir(sys.argv[1]):
            print("Not a valid path")
            return

    print("Path OK, continuing")

    #filename = date.strftime("%d_%b_%Y-%H_%M_%S")
    filename = "Default"
    path = sys.argv[1] + "\\" + str(filename) + ".xlsx"

    wb = openpyxl.Workbook()
    #ws = wb.active()
    wb.save(path)

    for elements in links:
        options = Options()
        options.add_argument('--headless=new')

        browser = webdriver.Chrome(options=options)
        browser.get(elements)

        time.sleep(5)

        html = browser.page_source

        getData(browser, path)

        browser.close()


precondition()
